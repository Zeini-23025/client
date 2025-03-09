
                           #PARTIE 1: IMPORTATION DES DONNÉES 
import pandas as pd
#Def table des matiers 
M=pd.read_excel('Données.xlsx',3)
Matieres=[]
for i in range(3,12):
  Matieres.append(M['Unnamed: 0'][i])

J=len(Matieres)

#Les charges
M1=pd.read_excel('Données.xlsx',3)
c=['Unnamed: 1','Unnamed: 2', 'Unnamed: 3', 'Unnamed: 4','Unnamed: 5', 'Unnamed: 6', 'Unnamed: 7', 'Unnamed: 8', 'Unnamed: 9','Unnamed: 10', 'Unnamed: 11', 'Unnamed: 12', 'Unnamed: 13','Unnamed: 14', 'Unnamed: 15', 'Unnamed: 16', 'Unnamed: 17','Unnamed: 18', 'Unnamed: 19', 'Unnamed: 20', 'Unnamed: 21','Unnamed: 22', 'Unnamed: 23', 'Unnamed: 24', 'Unnamed: 25','Unnamed: 26', 'Unnamed: 27', 'Unnamed: 28', 'Unnamed: 29','Unnamed: 30', 'Unnamed: 31', 'Unnamed: 32', 'Unnamed: 33']
Pcm=[]
Ptp=[]
Ptd=[]
G=len(c)
for j in range(3,12):
  CCM=[]
  CTP=[]
  CTD=[]
  for i in range(0,len(c),3):
    CCM.append(M1[c[i]][j])
    CTP.append(M1[c[i+1]][j])
    CTD.append(M1[c[i+2]][j])
  Pcm.append(CCM)
  Ptp.append(CTP)
  Ptd.append(CTD)

#Affectation des profs et creation des tableaux des profs par type d'enseignement
D=pd.read_excel('Données.xlsx',4)
B=pd.read_excel('Données.xlsx',1)
c=['Unnamed: 1','Unnamed: 2', 'Unnamed: 3', 'Unnamed: 4','Unnamed: 5', 'Unnamed: 6', 'Unnamed: 7', 'Unnamed: 8', 'Unnamed: 9','Unnamed: 10', 'Unnamed: 11', 'Unnamed: 12', 'Unnamed: 13','Unnamed: 14', 'Unnamed: 15', 'Unnamed: 16', 'Unnamed: 17','Unnamed: 18', 'Unnamed: 19', 'Unnamed: 20', 'Unnamed: 21','Unnamed: 22', 'Unnamed: 23', 'Unnamed: 24', 'Unnamed: 25','Unnamed: 26', 'Unnamed: 27', 'Unnamed: 28', 'Unnamed: 29','Unnamed: 30', 'Unnamed: 31', 'Unnamed: 32', 'Unnamed: 33']
profs=[]
for i in range(1,18):
  profs.append(B['Unnamed: 0'][i])
I=len(profs)
Ccm=[] 
Ctp=[]
Ctd=[]
for i in range(I):
  Ccm.append([])
  Ctp.append([])
  Ctd.append([])

t=0
ProCM=[]
ProTP=[]
ProTD=[]
for j in range(3,12):
  ProgCM=[]
  ProgTP=[]
  ProgTD=[]
  ProCM.append(ProgCM)
  ProTP.append(ProgTP)
  ProTD.append(ProgTD)
  g=0
  for k in range(0,len(c),3):
    ProgjCM=[]
    ProgjTP=[]
    ProgjTD=[]
    ProgCM.append(ProgjCM)
    ProgTP.append(ProgjTP)
    ProgTD.append(ProgjTD)
    for i in range(I):
      
       if D[c[k]][j]==profs[i]:   
          Ccm[i].append([g,t])
          ProgjCM.append(D[c[k]][j])
       if D[c[k+1]][j]==profs[i]:
          Ctp[i].append([g,t])
          ProgjTP.append(D[c[k+1]][j])
       if D[c[k+2]][j]==profs[i]:
          Ctd[i].append([g,t])
          ProgjTD.append(D[c[k+2]][j])
    g+=1
  t+=1
#Disponibilité des profs
Di=pd.read_excel('Données.xlsx',1)

a=['Lundi', 'Unnamed: 2', 'Unnamed: 3', 'Unnamed: 4','Unnamed: 5', 'Mardi', 'Unnamed: 7', 'Unnamed: 8', 'Unnamed: 9','Unnamed: 10', 'Mercredi ', 'Unnamed: 12', 'Unnamed: 13', 'Unnamed: 14','Unnamed: 15', 'Jeudi', 'Unnamed: 17', 'Unnamed: 18', 'Unnamed: 19','Unnamed: 20', 'Vendredi', 'Unnamed: 22', 'Unnamed: 23', 'Unnamed: 24','Unnamed: 25']
Dik=[]
for i in range(I):
  Dik.append([])
for i in a:
  for j in range(1,18):
      Dik[j-1].append(Di[i][j])

#Cheveauchement
D=pd.read_excel('Données.xlsx',2)

Groupes=['G1', 'G2', 'TP1', 'TP2', 'TP3', 'TP4', 'CNM', 'RSS','DSI_CM', 'DIS_TP1', 'DSI_TP2']
G=len(Groupes)
A=[]
for i in range(11):
  for j in range(11):
    if i<j:
      if D[Groupes[i]][j]==0:
        A.append([i,j])

S=4 # Nombre de salles.
K=25 # Nombre de créneaux
STP=2 # Nombre de salles de TP 

#indice des créneaux 
col=['B','C', 'D','E','F']
lin=['4','5','6','8','9']
ind=[]
for i in col:
  for j in lin:
    ind.append(i+j)
#ensemble des groupes 
E=[[0,2,3,6,7,8,9,10],[1,4,5,6,7,8,9,10]]

                                  #PARTIE 2: MODELE: IMPLEMENTATION ET RESOLUTION 

from ortools.linear_solver import pywraplp
import openpyxl
import pandas as pd

def LP_Emploi():
  solver = pywraplp.Solver.CreateSolver('CBC')
# Création des variables de décision
  X=[[[solver.IntVar(0, 1,'X'+str(g)+'_'+str(j)+'_'+str(k)) for k in range(K)]for j in range(J)] for g in range(G)] # CM
  Y=[[[solver.IntVar(0, 1,'Y'+str(g)+'_'+str(j)+'_'+str(k)) for k in range(K)]for j in range(J)] for g in range(G)] # TP
  Z=[[[solver.IntVar(0, 1,'Z'+str(g)+'_'+str(j)+'_'+str(k)) for k in range(K)]for j in range(J)] for g in range(G)] # TD
  
  # Contraintes1: Pour assurer que la charge de chaque matière sera dispensé complètement 
  for g in range(G):
      for j in range(J):
          solver.Add( sum(X[g][j][k] for k in range(K)) ==Pcm[j][g])     
          solver.Add( sum(Y[g][j][k] for k in range(K)) ==Ptp[j][g])
          solver.Add( sum(Z[g][j][k] for k in range(K)) ==Ptd[j][g])
 

 # Contraintes 2: Un groupe ne peut avoir qu’une seule séance dans un créneau précis
  for g in range(G):
    for k in range(K):
       solver.Add(sum(X[g][j][k]+Y[g][j][k]+Z[g][j][k] for j in range(J))<=1)

  # Contraintes 3: Disponibilité du local : le nombre de séances en parallèles ne doit pas dépasser le nombre de salles
  for k in range(K):
       solver.Add(sum(X[g][j][k]+Y[g][j][k]+Z[g][j][k] for j in range(J) for g in range(G))<=S)

  # Contraintes 4: Disponibilité du local : le nombre de séances de TP en parallèles ne doit pas dépasser le nombre de salles de TP
  for k in range(K):
       solver.Add(sum(Y[g][j][k] for j in range(J) for g in range(G))<=STP)

   #Contraintes 5: Quand un groupe est concerné par une séance, alors tous les sous groupes qui la constituent doivent l’étre et aucun d’eux ne peut se charger d’une autre séance

  for k in range(K):      
        for a in A:
            g1=a[0]
            g2=a[1]
            solver.Add(sum(X[g1][j][k]+Y[g1][j][k]+Z[g1][j][k]+X[g2][j][k] +Y[g2][j][k]+Z[g2][j][k] for j in range(J))<= 1)

#Contraintes 6:  La disponibilité de l’enseignent doit être respectée
  for i in range(I):
    for k in range(K):
      s=int(Dik[i][k])
      solver.Add( sum(X[h[0]][h[1]][k] for h in Ccm[i])+sum(Y[h[0]][h[1]][k] for h in Ctp[i])+sum(Z[h[0]][h[1]][k] for h in Ctd[i]) <= s) 
 
  
# Contraintes 7: La charge d’une matière soit dispensé par le prof associé  
  for i in range(I):
    for h in Ccm[i]:
         g=h[0]
         j=h[1]
         solver.Add( sum(X[g][j][k] for k in range(K)) ==Pcm[j][g]) 
  for i in range(I):
    for h in Ctp[i]:
         g=h[0]
         j=h[1]
         solver.Add( sum(Y[g][j][k] for k in range(K)) ==Ptp[j][g])
  for i in range(I):
    for h in Ctd[i]:
         g=h[0]
         j=h[1]
         solver.Add( sum(Z[g][j][k] for k in range(K)) ==Ptd[j][g])


  status = solver.Solve()

                              #PARTIE 3: EXPORTATION DES RESULTATS 
  print('Solution:')
  Ck=[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
  Ck1=[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]

  for j in range(J):
        for k in range(K):
            for g in range(G):

               if g in E[0]:                
                 if X[g][j][k].solution_value() !=0:
                     
                          wb=openpyxl.load_workbook('1emploi.xlsx')
                          ws=wb.active
                          ws[ind[k]].value = Groupes[g] +'/'+'CM: '+ Matieres[j]+'/ ' + str(ProCM[j][g])
                          wb.save('1emploi.xlsx')
                 if Y[g][j][k].solution_value() !=0:
                   if Ck[k]==0:
                   
                          wb=openpyxl.load_workbook('1emploi.xlsx')
                          ws=wb.active
                          ws[ind[k]].value =  Groupes[g]+': '+ Matieres[j]+'/ '+ str(ProTP[j][g])
                          wb.save('1emploi.xlsx')
                          Ck[k]+=1
                   else :
                   
                          wb=openpyxl.load_workbook('1emploi.xlsx')
                          ws=wb.active
                          ws[ind[k]].value = str(ws[ind[k]].value)+ Groupes[g]+': '+ Matieres[j]+str(ProTP[j][g])+ '//'
                          wb.save('1emploi.xlsx') 
                          Ck[k]+=1      
                 if Z[g][j][k].solution_value() ==1:
                   if Ck[k]==0:
                    
                          wb=openpyxl.load_workbook('1emploi.xlsx')
                          ws=wb.active
                          ws[ind[k]].value = Groupes[g]+': '+ Matieres[j]+'/ '+str(ProTD[j][g])
                          wb.save('1emploi.xlsx')
                          Ck[k]+=1
                   else:
                   
                          wb=openpyxl.load_workbook('1emploi.xlsx')
                          ws=wb.active
                          ws[ind[k]].value = str(ws[ind[k]].value)+ Groupes[g]+': '+ Matieres[j]+str(ProTD[j][g])+'//'
                          wb.save('1emploi.xlsx')
                          Ck[k]+=1
               if g in E[1]:   
                     
                 if X[g][j][k].solution_value() ==1:
                     
                          wb=openpyxl.load_workbook('2emploi.xlsx')
                          ws=wb.active
                          ws[ind[k]].value = Groupes[g] +'/'+'CM: '+ Matieres[j]+'/ ' + str(ProCM[j][g])
                          wb.save('2emploi.xlsx')
                 if Y[g][j][k].solution_value() ==1:
                   if Ck1[k]==0:
                  
                          wb=openpyxl.load_workbook('2emploi.xlsx')
                          ws=wb.active
                          ws[ind[k]].value =  Groupes[g]+': '+ Matieres[j]+'/ '+str(ProTP[j][g])
                          wb.save('2emploi.xlsx')
                          Ck1[k]+=1
                   else :
                
                          wb=openpyxl.load_workbook('2emploi.xlsx')
                          ws=wb.active
                          ws[ind[k]].value = str(ws[ind[k]].value)+ Groupes[g]+': '+ Matieres[j]+ str(ProTP[j][g])+'//'
                          wb.save('2emploi.xlsx') 
                          Ck1[k]+=1      
                 if Z[g][j][k].solution_value() ==1:
                   if Ck1[k]==0:
                  
                          wb=openpyxl.load_workbook('2emploi.xlsx')
                          ws=wb.active
                          ws[ind[k]].value = Groupes[g]+': '+ Matieres[j]+'/ '+str(ProTD[j][g])
                          wb.save('2emploi.xlsx')
                          Ck1[k]+=1 
                   else:
                  
                          wb=openpyxl.load_workbook('2emploi.xlsx')
                          ws=wb.active
                          ws[ind[k]].value = str(ws[ind[k]].value)+ Groupes[g]+': '+ Matieres[j]+str(ProTD[j][g])+'//'
                          wb.save('2emploi.xlsx')
                          Ck1[k]+=1 
  
  #Affichage des vars de décision 
  Xv=[]  
  Yv=[]  
  Zv=[]       
  for g in range(G):
   for j in range(J):
    for k in range(K):
      if X[g][j][k].solution_value()!=0 :
       Xv.append( 'X'+'_'+str(g)+'_'+str(j)+'_'+str(k))
      if Y[g][j][k].solution_value()!=0 :
       Yv.append( 'Y'+'_'+str(g)+'_'+str(j)+'_'+str(k))
      if Z[g][j][k].solution_value()!=0 :
       Zv.append( 'Z'+'_'+str(g)+'_'+str(j)+'_'+str(k))
  print(Xv,Yv,Zv)
LP_Emploi()

